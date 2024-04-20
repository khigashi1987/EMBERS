import os
import io
import glob
import logging
import openpyxl
import pandas as pd

class EXCELLoader():
    ###
    # PMC論文のexcelファイルを読み込み、テーブルを抽出するクラス
    ###
    def __init__(self):
        pass

    def is_merged_cell(self, sheet, row, col):
        """指定されたセルが結合されているかどうかを判断する"""
        for merged_range in sheet.merged_cells.ranges:
            if (row >= merged_range.min_row and row <= merged_range.max_row and
                col >= merged_range.min_col and col <= merged_range.max_col):
                return True
        return False

    def is_data_row(self, sheet, row):
        """実際のデータが含まれる行かどうかを判断する"""
        cell_count = 0
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row, column=col).value is not None and not self.is_merged_cell(sheet, row, col):
                cell_count += 1
        return cell_count >= 2

    def find_tables(self, sheet):
        tables = []
        current_table = None
        for row in range(1, sheet.max_row + 1):
            if self.is_data_row(sheet, row):
                if current_table is None:
                    current_table = [row, None]  # 新しいテーブルの開始
                elif current_table[1] is not None:
                    tables.append((sheet.title, tuple(current_table)))
                    current_table = [row, None]
            else:
                if current_table is not None:
                    current_table[1] = row - 1
                    tables.append((sheet.title, tuple(current_table)))
                    current_table = None

        if current_table is not None and current_table[1] is None:
            current_table[1] = sheet.max_row
            tables.append((sheet.title, tuple(current_table)))

        return tables

    def read_table_as_dataframe(self, file_path, sheet_name, start_row, end_row):
        return pd.read_excel(file_path, sheet_name=sheet_name, skiprows=start_row-1, nrows=end_row-start_row+1)

    def find_all_tables_in_workbook(self, file_path):
        workbook = openpyxl.load_workbook(file_path)
        all_tables_df = []

        for sheet in workbook:
            tables = self.find_tables(sheet)
            for table in tables:
                df = self.read_table_as_dataframe(file_path, table[0], table[1][0], table[1][1])
                # TEMP: 2行未満500行以上のテーブルはスキップ
                if len(df) > 500 or len(df) < 2:
                    continue
                # TEMP: 200列以上のテーブルはスキップ
                if len(df.columns) > 200:
                    continue

                all_tables_df += [df]

        return all_tables_df

    def analyze_excel(self, pmc_dir):
        article_excel_list = glob.glob(os.path.join(pmc_dir, '*.xlsx'))
        extracted_tables = []
        for excel_file in article_excel_list:
            ### TEMP
            if os.path.basename(excel_file).startswith('~'):
                continue

            # TEMP: 10MB以上のファイルはスキップ
            file_size = os.path.getsize(excel_file)
            size_in_mb = file_size / (1024 * 1024)
            if size_in_mb > 10:
                continue

            logging.info(f'\t\tanalyzing {excel_file}')
            extracted_tables += self.find_all_tables_in_workbook(excel_file)

            logging.info(f'\t\tanalyze done. {excel_file}')
        
        return extracted_tables

if __name__ == '__main__':

    import sys
    if len(sys.argv) != 2:
        print('Usage: python excelloader.py <PMC directory>')
        sys.exit()

    pmc_dir = sys.argv[1]

    loader = EXCELLoader()
    tables = loader.analyze_excel(pmc_dir)
    print(tables)
    print(f'extracted {len(tables)} tables.')

    for t in tables:
        print(t)
        print(t.shape)
        print(t.describe())